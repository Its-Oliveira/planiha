import pandas as pd
import openpyxl 
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill

df = pd.read_excel("orcamento.xlsx", header=None)
tamanho_plan = len(df.index)


def SyntaxBancos(df, l0, lf):
    for i in range(l0,lf):
        bc = df.loc[i,2]
        
        if bc == 'SINAPI-I' or bc == 'sinapi-i' or bc == 'sinapi-c' or bc == 'SINAPI-C' or bc =='sinapi' or bc=='SINAPI':
            df.loc[i,2] = "SINAPI"

        elif bc == 'SBC' or bc == 'sbc':
            df.loc[i,2] = 'SBC'

        elif bc=="CPOS" or bc=='cpos' or bc=='CDHU' or bc =='cdhu' or bc == 'CPOS/CDHU' or bc =='cpos/cdhu':
            df.loc[i,2] = 'CPOS'

        elif bc =="SICRO3" or bc == 'sicro3' or bc == 'SICRO' or bc == 'sicro':
            df.loc[i,2] = 'SICRO3' 

        elif bc == 'ORSE' or bc=='orse':
            df.loc[i,2] = 'ORSE'

        elif bc == 'sedop' or bc=='SEDOP':
            df.loc[i,2] = 'SEDOP'
            
        elif bc =='SEINFRA' or bc=='seinfra':
            df.loc[i,2] = 'SEINFRA'

        elif bc =='setop' or bc=='SETOP':
            df.loc[i,2] = 'SETOP'

        elif bc =='IOPES' or bc=='iopes':
            df.loc[i,2] = 'IOPES'

        elif bc =='SIURB' or bc=='siurb':
            df.loc[i,2] = 'SIURB'

        elif bc =='SIURB INFRA' or bc=='siurb infra' or bc=='SIURB infra' or bc=='siurb INFRA':
            df.loc[i,2] = 'SIURB INFRA'

        elif bc =='SUDECAP' or bc=='sudecap':
            df.loc[i,2] = 'SUDECAP'

        elif bc =='FDE' or bc=='fde':
            df.loc[i,2] = 'FDE'

        elif bc =='AGESUL' or bc=='agesul':
            df.loc[i,2] = 'AGESUL'

        elif bc =='FDE' or bc=='fde':
            df.loc[i,2] = 'FDE'
        
        elif bc =='AGETOP CIVIL' or bc=='agetop civil' or bc=='AGETOP civil' or bc=='agetop CIVIL':
            df.loc[i,2] = 'AGETOP CIVIL'

        elif bc =='AGETOP RODOVIARIA' or bc=='agetop rodoviaria' or bc=='AGETOP rodoviaria' or bc=='agetop RODOVIARIA':
            df.loc[i,2] = 'AGETOP RODOVIARIA'

        elif bc =='CAEMA' or bc=='caema':
            df.loc[i,2] = 'CAEMA'

        elif bc =='EMBASA' or bc=='embasa':
            df.loc[i,2] = 'EMBASA'

        elif bc =='CAERN' or bc=='caern':
            df.loc[i,2] = 'CAERN'

        elif bc =='COMPESA' or bc=='compesa':
            df.loc[i,2] = 'COMPESA'

        elif bc =='EMOP' or bc=='emop':
            df.loc[i,2] = 'EMOP'

        elif bc =='SCO' or bc=='sco':
            df.loc[i,2] = 'SCO'

    return df


def format_itemizacao(valor):
    try:
        partes = str(valor).split(".")
        partes_formatadas = [str(int(p)) for p in partes if p.strip() != ""]
        return ".".join(partes_formatadas)
    
    except ValueError:
        # Caso alguma parte não seja número, retorna o valor original
        return valor
    

def codes(df,l0,lf):
    for i in range(l0,lf):
        bc = df.loc[i,2]
        if bc == "SINAPI":
            code = str(df.loc[i,1])
            codel = (code.strip())
            df.loc[i,1] = codel
            if len(codel) == 5 or len(codel) == 6 or len(codel) == 8 or len(codel) == 4 or len(codel) == 9:
                df.loc[i,1] = codel
            else:
                df.loc[i,1] = "Erro, formato não reconhecido pela SINAPI"

        elif bc == "SBC":
            code = str(df.loc[i,1])
            codel = (code.strip())
            if len(codel) == 6:
                df.loc[i,1] = codel
            elif len(codel) == 1:
                codel = '00000'+codel
                df.loc[i,1] = codel
            elif len(codel) == 2:
                codel = '0000'+codel
                df.loc[i,1] = codel
            elif len(codel) == 3:
                codel = '000'+codel
                df.loc[i,1] = codel
            elif len(codel) == 4:
                codel = '00'+codel
                df.loc[i,1] = codel
            elif len(codel) == 5:
                codel = '0'+codel
                df.loc[i,1] = codel
            else:
                df.loc[i,1] = "Erro, formato não reconhecido pela SBC"  


        elif bc == "CPOS":
            code = str(df.loc[i,1])
            codel = (code.strip())
            df.loc[i,1] = codel
            if len(codel) == 9 or len(codel) == 15:
                df.loc[i,1] = codel
            else:
                df.loc[i,1] = "Erro, formato não reconhecido pela CPOS"

        elif bc == "SICRO":
            code = str(df.loc[i,1])
            codel = (code.strip())
            df.loc[i,1] = codel
            if len(codel) == 7 or len(codel) == 5:
                df.loc[i,1] = codel
            else:
                df.loc[i,1] = "Erro, formato não reconhecido pela SICRO"

        elif bc == "SETOP":
            code = str(df.loc[i,1])
            codel = (code.strip())
            df.loc[i,1] = codel
            if len(codel) == 8 or len(codel) == 11:
                df.loc[i,1] = codel
            else:
                df.loc[i,1] = "Erro, formato não reconhecido pela SETOP"

        elif bc == "IOPES":
            code = str(df.loc[i,1])
            codel = (code.strip())
            df.loc[i,1] = codel
            if len(codel) == 6:
                df.loc[i,1] = codel
            else:
                df.loc[i,1] = "Erro, formato não reconhecido pela IOPES"

        elif bc == "SIURB":
            code = str(df.loc[i,1])
            codel = (code.strip())
            df.loc[i,1] = codel
            if len(codel) == 7 or len(codel) == 5 or len(codel) == 8:
                df.loc[i,1] = codel
            else:
                df.loc[i,1] = "Erro, formato não reconhecido pela SIURB"

        elif bc == "SIRUB INFRA":
            code = str(df.loc[i,1])
            codel = (code.strip())
            df.loc[i,1] = codel
            if len(codel) == 7 or len(codel) == 5:
                df.loc[i,1] = codel
            else:
                df.loc[i,1] = "Erro, formato não reconhecido pela SIURB INFRA"

        elif bc == "SUDECAP":
            code = str(df.loc[i,1])
            codel = (code.strip())
            df.loc[i,1] = codel
            if len(codel) == 8:
                df.loc[i,1] = codel
            else:
                df.loc[i,1] = "Erro, formato não reconhecido pela SUDECAP"

        elif bc == "FDE":
            code = str(df.loc[i,1])
            codel = (code.strip())
            df.loc[i,1] = codel
            if len(codel) == 7 or len(codel) == 9:
                df.loc[i,1] = codel
            else:
                df.loc[i,1] = "Erro, formato não reconhecido pela FDE"

        elif bc == "EMOP":
            code = str(df.loc[i,1])
            codel = (code.strip())
            df.loc[i,1] = codel
            if len(codel) == 13 or len(codel) == 5:
                df.loc[i,1] = codel
            else:
                df.loc[i,1] = "Erro, formato não reconhecido pela EMOP"

        elif bc == "SCO":
            code = str(df.loc[i,1])
            codel = (code.strip())
            df.loc[i,1] = codel
            if len(codel) == 13 or len(codel) == 9:
                df.loc[i,1] = codel
            else:
                df.loc[i,1] = "Erro, formato não reconhecido pela SCO"
                print(len(codel))

        elif bc == "ORSE":
            code = str(df.loc[i,1])
            codel = (code.strip())
            df.loc[i,1] = codel
            if len(codel) == 4 or len(codel) == 5 or len(codel) == 15 or len(codel) == 3 or len(codel) == 2 :
                df.loc[i,1] = codel
            else:
                df.loc[i,1] = "Erro, formato não reconhecido pela ORSE"

        elif bc == "SEINFRA":
            code = str(df.loc[i,1])
            codel = (code.strip())
            df.loc[i,1] = codel
            if len(codel) == 5:
                df.loc[i,1] = codel
            else:
                df.loc[i,1] = "Erro, formato não reconhecido pela SEINFRA"

        elif bc == "CAEMA":
            code = str(df.loc[i,1])
            codel = (code.strip())
            df.loc[i,1] = codel
            if len(codel) == 6 or len(codel) == 10:
                df.loc[i,1] = codel
            else:
                df.loc[i,1] = "Erro, formato não reconhecido pela CAEMA"

        elif bc == "EMBASA":
            code = str(df.loc[i,1])
            codel = (code.strip())
            df.loc[i,1] = codel
            if len(codel) == 8 or len(codel) == 10:
                df.loc[i,1] = codel
            else:
                df.loc[i,1] = "Erro, formato não reconhecido pela EMBASA"

        elif bc == "CAERN":
            code = str(df.loc[i,1])
            codel = (code.strip())
            df.loc[i,1] = codel
            if len(codel) == 3 or len(codel) == 4 or len(codel) == 5 or len(codel) == 6 or len(codel) == 7:
                df.loc[i,1] = codel
            else:
                df.loc[i,1] = "Erro, formato não reconhecido pela CAERN"

        elif bc == "COMPESA":
            code = str(df.loc[i,1])
            codel = (code.strip())
            df.loc[i,1] = codel
            if len(codel) == 9 or len(codel) == 13 or len(codel) == 15:
                df.loc[i,1] = codel
            else:
                df.loc[i,1] = "Erro, formato não reconhecido pela COMPESA"

        elif bc == "AGESUL":
            code = str(df.loc[i,1])
            codel = (code.strip())
            df.loc[i,1] = codel
            if len(codel) == 4 or len(codel) == 10:
                df.loc[i,1] = codel
            else:
                df.loc[i,1] = "Erro, formato não reconhecido pela AGESUL"

        elif bc == "AGETOP CIVIL":
            code = str(df.loc[i,1])
            codel = (code.strip())
            df.loc[i,1] = codel
            if len(codel) == 6 or len(codel) == 4:
                df.loc[i,1] = codel
            else:
                df.loc[i,1] = "Erro, formato não reconhecido pela AGETOP CIVIL"

        elif bc == "AGETOP RODOVIARIA":
            code = str(df.loc[i,1])
            codel = (code.strip())
            df.loc[i,1] = codel
            if len(codel) == 5:
                df.loc[i,1] = codel
            else:
                df.loc[i,1] = "Erro, formato não reconhecido pela AGETOP RODOVIARIA"

        elif bc == "SEDOP":
            code = str(df.loc[i,1])
            codel = (code.strip())
            df.loc[i,1] = codel
            if len(codel) == 8 or len(codel) == 6:
                df.loc[i,1] = codel
            else:
                df.loc[i,1] = "Erro, formato não reconhecido pela SEDOP"

        elif bc == "DERPR":
            code = str(df.loc[i,1])
            codel = (code.strip())
            df.loc[i,1] = codel
            if len(codel) == 6:
                df.loc[i,1] = codel
            else:
                df.loc[i,1] = "Erro, formato não reconhecido pela DERPR"


        

df.to_excel('Planilha Ajustada.xlsx')

   

def format_plan_ajustada(ws,wb):
    wb = openpyxl.load_workbook('Planilha Ajustada.xlsx')
    ws = wb["Sheet1"]

    for col in ws.columns:
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 80
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 15
    ws.row_dimensions[1].height = 30.0
    ws.row_dimensions[2].height = 30.0


    ws = wb.active
    fonte_negrito = Font(bold= True, size=11,name='Arial')
    alinhado_centro = Alignment(horizontal='left',vertical='center')
    alinhado_esquerda = Alignment(horizontal="left",vertical='center')
    borda_fina = Side(border_style="thin", color="000000")

    for row in ws['A1':"I1"]: #deixando cabeçalho em negrito
        for cell in row:
            cell.alignment = alinhado_centro
            cell.font = fonte_negrito
            cell.border = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)
        

    for row in ws.iter_rows(): #formatar alinhamento e bordas dos itens do orçamento
        for cell in row:
            cell.alignment = alinhado_esquerda
            cell.border = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)

    for i in range(1,ws.max_row+1): #formatar tamanho da linha
        ws.row_dimensions[i].height = 30.0


    wb.save('Planilha Ajustada.xlsx')